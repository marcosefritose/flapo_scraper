# ENV
import os
from django.db import models
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from django.core.management.base import BaseCommand
from scraping.models import Review, Platform, Text
import datetime


class Command(BaseCommand):
    help = "Collect review from Indeed"

    def handle(self, *args, **options):

        def get_header_row_values(row):
          header_row = []

          for column_name in row:
              header_row.append(column_name.value)

          return header_row

        def parse_author_status(status_string):
            if status_string == 'Ex-Job':
                return False
            elif status_string == 'Aktueller Job':
                return True
            else:
                return None

        platform, created = Platform.objects.get_or_create(
            title='Kununu',
            defaults={'last_scraped': None}
        )

        new_review_count = 0

        # ToDo add ftp access to xls file
        path = os.getenv('KUNUNU_EXPORT_PATH')
        export_object = openpyxl.load_workbook(path)

        rating_sheet = export_object.worksheets[1]
        reviews_sheet = export_object.worksheets[4]

        review_header_row = get_header_row_values(rating_sheet[5])
        text_header_row = get_header_row_values(reviews_sheet[5])

        for row in rating_sheet.iter_rows(min_row=6):
          # date = datetime.datetime.strptime(row[0].value, "%d.%m.%Y")
          date = row[0].value
          rating = float(row[1].value)
          title = row[2].value

          author_status = parse_author_status(row[16].value)
          author_role = row[20].value
          author_location = row[18].value

          text_ref = row[22].value
          
          try:
            # Try to create new or match with existing review based on review URL
            review, created = Review.objects.update_or_create(
                #ToDo find unique identifier
                title=title,
                defaults={
                    'platform': platform,
                    'creation_date': date,
                    'total_rating_score': rating,
                    'author_status': author_status,
                    'author_role': author_role,
                    'author_location': author_location
                }
            )
            if created:
              new_review_count = new_review_count + 1

              if text_ref is not None:
                for review_row in reviews_sheet.iter_rows(min_row=6):
                  
                  if review_row[25].value == text_ref:
                      for index, review_text in enumerate(review_row[3:20]):

                        if review_text.value is not None:
                          text = Text.objects.create(
                            review = review,
                            text_type = text_header_row[index+3],
                            content = review_text.value
                          )

          except Exception as e:
            print(e)

        platform.last_scraped = datetime.date.today()
        platform.save()

        self.stdout.write(f"Kununu review import complete. {new_review_count} new reviews added.")

